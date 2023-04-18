import os
from openpyxl import load_workbook, Workbook

def get_all_dirs(path):
    dirs = []
    files = []
    for (r, d, f) in os.walk(path):
        dirs.append(d)
        files.append(f)
    return dirs

def find_files(path):
    files = list(os.walk(path))
    return files

def group_house_sheets(files_list, dir_path):
    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet.title = "Total_Consumers"

    new_file_row = 0
    cpe = ""
    new_sheet["A1"] = "Data"
    new_sheet["B1"] = "Hora"
    new_sheet["C1"] = "Consumo Registado"

    for f in files_list:
        if f.endswith(".xlsx"):
            wb = load_workbook(filename=dir_path+"/"+f)
            sheet = wb.active
            current_row = 0
            startCollecting = False
            for row in sheet:
                for cell in row:
                    if(startCollecting):
                        new_sheet.cell(row=current_row+new_file_row+1, column=cell.col_idx, value=cell.value)
                    if(cell.value == "Data"):
                        startCollecting = True
                        break
                    if(isinstance(cell.value, str) and cell.value.startswith("PT")):
                        cpe = cell.value
                if(startCollecting):
                    current_row += 1
        new_file_row += current_row
    new_wb.save("../../datasets/Our_Dataset/" + cpe + ".xlsx")
    
def rename_sheets(files_list, dir_path):
    for f in files_list:
        if f.endswith(".xlsx"):
            filename = dir_path + "/" + f
            wb = load_workbook(filename=filename)
            sheet = wb.active
            date = sheet["A12"].value
            wb.close()
            new_name = dir_path + "/" + date[:4] + date[5:7] + ".xlsx"
            print(new_name)
            if(not os.path.exists(new_name)):
                os.rename(filename, new_name)

def organize_by_datetime(files_list):
    files_list = list(filter(lambda f: f.endswith(".xlsx") and f.startswith("20"), files_list))
    files_list.sort()
    return files_list

def main():
    path = "../../datasets/Curvas_de_Consumo/"

    dirs = get_all_dirs(path)[0]
    for dir in dirs:
        if len(dir) > 0:
            dir_path = path + "/" + dir
            files_list = find_files(dir_path)[0][2]
            print(files_list)
            rename_sheets(files_list, dir_path)
            new_files_list = find_files(dir_path)[0][2]
            new_files_list = organize_by_datetime(new_files_list)
            group_house_sheets(new_files_list, dir_path)

if __name__ == "__main__":
    main()
